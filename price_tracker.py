from pycoingecko import CoinGeckoAPI
from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

cg = CoinGeckoAPI()


filename = 'crypto_names.txt'


def names():

    yn = input("Are you using it for the first time? => ").lower()
    print(' ')

    filename = 'crypto_names.txt'


    filename = 'crypto_names.txt'

    if "no" in yn:
        
        lines = []
        with open(filename, 'r') as file: 
            for line in file: 
                lines.append(line.strip())


        for i in range(0, len(lines)):
                
            names = lines[i]
            print(f"List of items: {names} \n")


        cryptos = ','.join(lines)
        prices = cg.get_price(ids=cryptos, vs_currencies='usd')

        workbook = load_workbook("Price_Data.xlsx")
        sheet = workbook.active

        next_row = sheet.max_row + 1

        heading = "Names" 
        date_heading = "Date"

        bold_font = Font(bold=True)
        sheet["A1"] = heading 
        sheet["A1"].font = bold_font

        sheet["B1"] = date_heading 
        sheet["B1"].font = bold_font

        sheet["C1"] = "Prices"
        sheet["C1"].font = bold_font

        items = lines

        for index, item in enumerate(items, start=next_row): 
            sheet[f"A{index}"] = item
            sheet[f"B{index}"] = datetime.now().strftime("%Y-%m-%d")
            sheet[f"C{index}"] = prices.get(item, {}).get('usd', 'N/A') 


        workbook.save("Price_Data.xlsx")

        print("YOUR CRYPTO PRICES SHEET UPDATED")
        print(" ")

        
    elif "yes" in yn:
        user_input = input("Enter crypto names separated by commas: ")
        print(" ")
        print("Sheet created \n")
        items = user_input.split(',')
    
        with open('crypto_names.txt', 'w') as file:
    
            for item in items:
                file.write(item.strip() + '\n')



        lines = []
        with open(filename, 'r') as file: 
            for line in file: 
                lines.append(line.strip())

        for i in range(0, len(lines)):
                
            names = lines[i]
            print(f"List of items: {names} \n")

        cryptos = ','.join(lines)
        prices = cg.get_price(ids=cryptos, vs_currencies='usd')

    
        
        workbook = Workbook() 
        sheet = workbook.active

        heading = "Names" 
        date_heading = "Date"

        bold_font = Font(bold=True)
        sheet["A1"] = heading 
        sheet["A1"].font = bold_font

        sheet["B1"] = date_heading 
        sheet["B1"].font = bold_font

        sheet["C1"] = "Prices"
        sheet["C1"].font = bold_font

        items = lines

        for index, item in enumerate(items, start=2): 
            sheet[f"A{index}"] = item
            sheet[f"B{index}"] = datetime.now().strftime("%Y-%m-%d")
            sheet[f"C{index}"] = prices.get(item, {}).get('usd', 'N/A') 

        workbook.save("Price_Data.xlsx")

        print("SHEET CRAETED SUCCESFULLY")
        print(" ")

    else:
        print("Unable to record data")

names()

#CODE SOURCED FROM Github=>DivyanshAtray

